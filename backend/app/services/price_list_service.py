"""
================================================================================
Marketplace ERP - Customer Price List Service
================================================================================
Version: 1.0.0
Created: 2025-12-11

Service for managing customer price lists, items, Excel import/export,
price resolution, and history tracking.
================================================================================
"""

from typing import List, Dict, Optional, Tuple
from fastapi import HTTPException, status, UploadFile
from datetime import date, datetime
from decimal import Decimal
import logging
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import fetch_one, fetch_all, execute_query

logger = logging.getLogger(__name__)


# ============================================================================
# PRICE LIST CRUD OPERATIONS
# ============================================================================

async def get_price_lists(
    status_filter: Optional[str] = None,
    date_filter: Optional[date] = None,
    page: int = 1,
    limit: int = 50
) -> Dict:
    """
    Get all price lists with optional filters
    
    Args:
        status_filter: Filter by status (active, expired, upcoming)
        date_filter: Get lists active on specific date
        page: Page number
        limit: Results per page
        
    Returns:
        Dict with price_lists, total, page, limit
    """
    where_conditions = []
    params = []
    param_count = 1
    
    # Filter by status
    if status_filter:
        today = date.today()
        if status_filter == "active":
            where_conditions.append(
                f"(valid_from <= ${param_count} AND (valid_to IS NULL OR valid_to >= ${param_count}) AND is_active = TRUE)"
            )
            params.append(today)
            param_count += 1
        elif status_filter == "expired":
            where_conditions.append(f"valid_to < ${param_count}")
            params.append(today)
            param_count += 1
        elif status_filter == "upcoming":
            where_conditions.append(f"valid_from > ${param_count}")
            params.append(today)
            param_count += 1
    
    # Filter by specific date
    if date_filter:
        where_conditions.append(
            f"(valid_from <= ${param_count} AND (valid_to IS NULL OR valid_to >= ${param_count}))"
        )
        params.append(date_filter)
        param_count += 1
    
    where_clause = f"WHERE {' AND '.join(where_conditions)}" if where_conditions else ""
    
    # Get total count
    count_query = f"""
        SELECT COUNT(*) as total
        FROM customer_price_lists
        {where_clause}
    """
    count_result = await fetch_one(count_query, *params)
    total = count_result["total"] if count_result else 0
    
    # Get paginated results
    offset = (page - 1) * limit
    query = f"""
        SELECT 
            cpl.*,
            (SELECT COUNT(*) FROM price_list_items WHERE price_list_id = cpl.id) as items_count,
            (SELECT COUNT(*) FROM zoho_customers WHERE price_list_id = cpl.id) as customers_count
        FROM customer_price_lists cpl
        {where_clause}
        ORDER BY valid_from DESC, created_at DESC
        LIMIT ${param_count} OFFSET ${param_count + 1}
    """
    params.extend([limit, offset])
    
    price_lists_raw = await fetch_all(query, *params)
    
    # Add status field
    price_lists = []
    today = date.today()
    for pl in price_lists_raw:
        pl_dict = dict(pl)
        # Convert UUID to string (ALWAYS convert, even if None)
        pl_dict['created_by'] = str(pl_dict.get('created_by')) if pl_dict.get('created_by') is not None else None
        # Determine status
        if pl_dict['valid_from'] > today:
            pl_dict['status'] = 'upcoming'
        elif pl_dict['valid_to'] and pl_dict['valid_to'] < today:
            pl_dict['status'] = 'expired'
        elif pl_dict['is_active']:
            pl_dict['status'] = 'active'
        else:
            pl_dict['status'] = 'inactive'
        price_lists.append(pl_dict)
    
    return {
        "price_lists": price_lists,
        "total": total,
        "page": page,
        "limit": limit
    }


async def get_price_list_by_id(price_list_id: int) -> Dict:
    """Get price list by ID with counts"""
    query = """
        SELECT 
            cpl.*,
            (SELECT COUNT(*) FROM price_list_items WHERE price_list_id = cpl.id) as items_count,
            (SELECT COUNT(*) FROM zoho_customers WHERE price_list_id = cpl.id) as customers_count
        FROM customer_price_lists cpl
        WHERE cpl.id = $1
    """
    result = await fetch_one(query, price_list_id)
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Price list with ID {price_list_id} not found"
        )
    
    pl_dict = dict(result)
    
    # Convert UUID to string for Pydantic
    pl_dict['created_by'] = str(pl_dict.get('created_by')) if pl_dict.get('created_by') is not None else None
    
    # Add status
    today = date.today()
    if pl_dict['valid_from'] > today:
        pl_dict['status'] = 'upcoming'
    elif pl_dict['valid_to'] and pl_dict['valid_to'] < today:
        pl_dict['status'] = 'expired'
    elif pl_dict['is_active']:
        pl_dict['status'] = 'active'
    else:
        pl_dict['status'] = 'inactive'
    
    return pl_dict


async def create_price_list(data: Dict, created_by: str) -> Dict:
    """Create new price list"""
    # Check for duplicate name
    existing = await fetch_one(
        "SELECT id FROM customer_price_lists WHERE price_list_name = $1",
        data['price_list_name']
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Price list with name '{data['price_list_name']}' already exists"
        )
    
    query = """
        INSERT INTO customer_price_lists 
        (price_list_name, description, valid_from, valid_to, is_active, created_by)
        VALUES ($1, $2, $3, $4, $5, $6)
        RETURNING *
    """
    
    result = await fetch_one(
        query,
        data['price_list_name'],
        data.get('description'),
        data['valid_from'],
        data.get('valid_to'),
        data.get('is_active', True),
        created_by
    )
    
    result = dict(result)
    
    # Convert UUID to string for Pydantic
    result['created_by'] = str(created_by) if created_by is not None else None
    
    # Add counts (will be 0 for a new price list)
    result['items_count'] = 0
    result['customers_count'] = 0
    
    # Determine status
    today = date.today()
    if result['valid_from'] > today:
        result['status'] = 'upcoming'
    elif result['valid_to'] and result['valid_to'] < today:
        result['status'] = 'expired'
    elif result['is_active']:
        result['status'] = 'active'
    else:
        result['status'] = 'inactive'
    
    return result


async def update_price_list(price_list_id: int, data: Dict, updated_by: str) -> Dict:
    """Update price list"""
    # Check exists
    existing = await get_price_list_by_id(price_list_id)
    
    # Check name uniqueness if changing name
    if 'price_list_name' in data and data['price_list_name'] != existing['price_list_name']:
        name_check = await fetch_one(
            "SELECT id FROM customer_price_lists WHERE price_list_name = $1 AND id != $2",
            data['price_list_name'],
            price_list_id
        )
        if name_check:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Price list with name '{data['price_list_name']}' already exists"
            )
    
    # Build update query dynamically
    update_fields = []
    params = []
    param_count = 1
    
    for field in ['price_list_name', 'description', 'valid_from', 'valid_to', 'is_active']:
        if field in data:
            update_fields.append(f"{field} = ${param_count}")
            params.append(data[field])
            param_count += 1
    
    if not update_fields:
        return existing
    
    # Add updated_at
    update_fields.append(f"updated_at = NOW()")
    
    # Add ID to params
    params.append(price_list_id)
    
    query = f"""
        UPDATE customer_price_lists
        SET {', '.join(update_fields)}
        WHERE id = ${param_count}
    """
    
    await execute_query(query, *params)
    
    return await get_price_list_by_id(price_list_id)


async def delete_price_list(price_list_id: int) -> None:
    """
    Delete price list - validates no customers assigned
    """
    # Check if any customers assigned
    customers = await fetch_all(
        """
        SELECT zc.id, zc.company_name 
        FROM zoho_customers zc
        WHERE zc.price_list_id = $1
        LIMIT 5
        """,
        price_list_id
    )
    
    if customers:
        customer_names = [c['company_name'] for c in customers]
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete price list. {len(customers)} customers assigned. "
                   f"First few: {', '.join(customer_names[:3])}. "
                   f"Please remove all customer assignments first."
        )
    
    # Delete price list (cascade will delete items and history)
    result = await execute_query(
        "DELETE FROM customer_price_lists WHERE id = $1",
        price_list_id
    )
    
    if result == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Price list with ID {price_list_id} not found"
        )


async def duplicate_price_list(
    price_list_id: int,
    new_name: str,
    copy_items: bool,
    valid_from: Optional[date],
    valid_to: Optional[date],
    created_by: str
) -> Dict:
    """Duplicate price list with optional items copy"""
    # Get original
    original = await get_price_list_by_id(price_list_id)
    
    # Create new price list
    new_data = {
        'price_list_name': new_name,
        'description': original.get('description'),
        'valid_from': valid_from or original['valid_from'],
        'valid_to': valid_to,
        'is_active': True
    }
    
    new_pl = await create_price_list(new_data, created_by)
    
    # Copy items if requested
    if copy_items:
        items = await fetch_all(
            "SELECT item_id, price, notes FROM price_list_items WHERE price_list_id = $1",
            price_list_id
        )
        
        for item in items:
            await execute_query(
                """
                INSERT INTO price_list_items (price_list_id, item_id, price, notes)
                VALUES ($1, $2, $3, $4)
                """,
                new_pl['id'],
                item['item_id'],
                item['price'],
                item.get('notes')
            )
    
    return await get_price_list_by_id(new_pl['id'])


# ============================================================================
# PRICE LIST ITEMS MANAGEMENT
# ============================================================================

async def get_price_list_items(price_list_id: int) -> Dict:
    """Get all items in a price list"""
    # Verify price list exists
    await get_price_list_by_id(price_list_id)
    
    query = """
        SELECT 
            pli.*,
            zi.name as item_name,
            zi.sku as item_sku
        FROM price_list_items pli
        JOIN zoho_items zi ON zi.id = pli.item_id
        WHERE pli.price_list_id = $1
        ORDER BY zi.name
    """
    
    items = await fetch_all(query, price_list_id)
    
    return {
        "items": [dict(i) for i in items],
        "total": len(items)
    }


async def add_or_update_price_list_item(
    price_list_id: int,
    item_id: int,
    price: Decimal,
    notes: Optional[str] = None
) -> Dict:
    """Add or update single item in price list"""
    # Verify price list exists
    await get_price_list_by_id(price_list_id)
    
    # Verify item exists
    item = await fetch_one("SELECT id, name FROM zoho_items WHERE id = $1", item_id)
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Item with ID {item_id} not found"
        )
    
    # Upsert
    query = """
        INSERT INTO price_list_items (price_list_id, item_id, price, notes)
        VALUES ($1, $2, $3, $4)
        ON CONFLICT (price_list_id, item_id)
        DO UPDATE SET price = $3, notes = $4, updated_at = NOW()
        RETURNING id
    """
    
    result = await fetch_one(query, price_list_id, item_id, price, notes)
    
    # Get full item details
    item_result = await fetch_one(
        """
        SELECT 
            pli.*,
            zi.name as item_name,
            zi.sku as item_sku
        FROM price_list_items pli
        JOIN zoho_items zi ON zi.id = pli.item_id
        WHERE pli.id = $1
        """,
        result['id']
    )
    
    return dict(item_result)


async def bulk_add_or_update_items(
    price_list_id: int,
    items: List[Dict]
) -> Dict:
    """Bulk add or update items"""
    # Verify price list
    await get_price_list_by_id(price_list_id)
    
    added = 0
    updated = 0
    errors = []
    
    for item_data in items:
        try:
            # Check if exists
            existing = await fetch_one(
                "SELECT id FROM price_list_items WHERE price_list_id = $1 AND item_id = $2",
                price_list_id,
                item_data['item_id']
            )
            
            await add_or_update_price_list_item(
                price_list_id,
                item_data['item_id'],
                item_data['price'],
                item_data.get('notes')
            )
            
            if existing:
                updated += 1
            else:
                added += 1
                
        except Exception as e:
            errors.append({
                "item_id": item_data['item_id'],
                "error": str(e)
            })
    
    return {
        "added": added,
        "updated": updated,
        "errors": errors,
        "total_processed": len(items)
    }


async def delete_price_list_item(price_list_id: int, item_id: int) -> None:
    """Remove item from price list"""
    result = await execute_query(
        "DELETE FROM price_list_items WHERE price_list_id = $1 AND item_id = $2",
        price_list_id,
        item_id
    )
    
    if result == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Item not found in price list"
        )


# ============================================================================
# EXCEL IMPORT/EXPORT
# ============================================================================

async def generate_excel_template() -> bytes:
    """Generate Excel template for price list import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Price List Template"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["SKU", "Item Name (Read-Only)", "Price (INR)", "Notes (Optional)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sample row
    ws.cell(row=2, column=1, value="SKU001")
    ws.cell(row=2, column=2, value="Example Item Name")
    ws.cell(row=2, column=3, value=100.00)
    ws.cell(row=2, column=4, value="Sample notes")
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


async def import_from_excel(price_list_id: int, file: UploadFile) -> Dict:
    """Import price list items from Excel file"""
    # Verify price list
    await get_price_list_by_id(price_list_id)
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        items_imported = 0
        items_updated = 0
        errors = []
        
        # Skip header row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            sku = str(row[0]).strip()
            
            try:
                price = float(row[2]) if row[2] else None
                notes = str(row[3]) if row[3] else None
                
                if not price or price <= 0:
                    errors.append({
                        "row_number": row_num,
                        "sku": sku,
                        "error": "Price must be greater than 0"
                    })
                    continue
                
                # Find item by SKU
                item = await fetch_one(
                    "SELECT id FROM zoho_items WHERE sku = $1",
                    sku
                )
                
                if not item:
                    errors.append({
                        "row_number": row_num,
                        "sku": sku,
                        "error": "SKU not found in items database"
                    })
                    continue
                
                # Check if exists
                existing = await fetch_one(
                    "SELECT id FROM price_list_items WHERE price_list_id = $1 AND item_id = $2",
                    price_list_id,
                    item['id']
                )
                
                # Add or update
                await add_or_update_price_list_item(
                    price_list_id,
                    item['id'],
                    Decimal(str(price)),
                    notes
                )
                
                if existing:
                    items_updated += 1
                else:
                    items_imported += 1
                    
            except Exception as e:
                errors.append({
                    "row_number": row_num,
                    "sku": sku,
                    "error": str(e)
                })
        
        return {
            "success": True,
            "items_imported": items_imported,
            "items_updated": items_updated,
            "items_failed": len(errors),
            "errors": errors,
            "message": f"Imported {items_imported}, Updated {items_updated}, Failed {len(errors)}"
        }
        
    except Exception as e:
        logger.error(f"Excel import error: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}"
        )


async def export_to_excel(price_list_id: int) -> bytes:
    """Export price list items to Excel"""
    # Get price list info
    price_list = await get_price_list_by_id(price_list_id)
    
    # Get items
    items_response = await get_price_list_items(price_list_id)
    items = items_response['items']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Price List"
    
    # Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["SKU", "Item Name", "Price (INR)", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item['item_sku'])
        ws.cell(row=row_num, column=2, value=item['item_name'])
        ws.cell(row=row_num, column=3, value=float(item['price']))
        ws.cell(row=row_num, column=4, value=item.get('notes', ''))
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    
    # Save
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()


# ============================================================================
# CUSTOMER ASSIGNMENT
# ============================================================================

async def get_assigned_customers(price_list_id: int) -> Dict:
    """Get all customers assigned to a price list"""
    # Verify price list exists
    price_list = await get_price_list_by_id(price_list_id)
    
    query = """
        SELECT 
            zc.id as customer_id,
            zc.company_name,
            zc.contact_name,
            zc.price_list_id
        FROM zoho_customers zc
        WHERE zc.price_list_id = $1
        ORDER BY zc.company_name
    """
    
    customers = await fetch_all(query, price_list_id)
    
    # Add price list status for each
    today = date.today()
    customers_list = []
    for c in customers:
        c_dict = dict(c)
        c_dict['price_list_name'] = price_list['price_list_name']
        
        if price_list['valid_from'] > today:
            c_dict['price_list_status'] = 'upcoming'
        elif price_list['valid_to'] and price_list['valid_to'] < today:
            c_dict['price_list_status'] = 'expired'
        elif price_list['is_active']:
            c_dict['price_list_status'] = 'active'
        else:
            c_dict['price_list_status'] = 'inactive'
        
        customers_list.append(c_dict)
    
    return {
        "price_list_id": price_list_id,
        "price_list_name": price_list['price_list_name'],
        "customers": customers_list,
        "total": len(customers_list)
    }


# ============================================================================
# PRICE RESOLUTION
# ============================================================================

async def resolve_customer_item_price(
    customer_id: int,
    item_id: int,
    date_for: Optional[date] = None
) -> Dict:
    """
    Resolve price for customer + item combination
    
    Priority:
    1. Valid price list price
    2. Zoho default price (from zoho_items.rate)
    """
    if not date_for:
        date_for = date.today()
    
    # Get customer with price list
    customer = await fetch_one(
        "SELECT id, price_list_id FROM zoho_customers WHERE id = $1",
        customer_id
    )
    
    if not customer:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Customer with ID {customer_id} not found"
        )
    
    # Get item
    item = await fetch_one(
        "SELECT id, rate FROM zoho_items WHERE id = $1",
        item_id
    )
    
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Item with ID {item_id} not found"
        )
    
    result = {
        "customer_id": customer_id,
        "item_id": item_id,
        "date_resolved_for": date_for
    }
    
    # Check if customer has price list assigned
    if customer['price_list_id']:
        # Get price list
        price_list = await fetch_one(
            """
            SELECT id, price_list_name, valid_from, valid_to, is_active
            FROM customer_price_lists
            WHERE id = $1
            """,
            customer['price_list_id']
        )
        
        if price_list:
            # Check if valid for date
            is_valid = (
                price_list['is_active'] and
                price_list['valid_from'] <= date_for and
                (price_list['valid_to'] is None or price_list['valid_to'] >= date_for)
            )
            
            result['price_list_id'] = price_list['id']
            result['price_list_name'] = price_list['price_list_name']
            result['is_price_list_active'] = is_valid
            
            if is_valid:
                # Look up item price in price list
                price_list_item = await fetch_one(
                    "SELECT price FROM price_list_items WHERE price_list_id = $1 AND item_id = $2",
                    price_list['id'],
                    item_id
                )
                
                if price_list_item:
                    result['price'] = price_list_item['price']
                    result['source'] = 'price_list'
                    return result
    
    # Fallback to Zoho default
    result['price'] = item['rate'] or Decimal('0')
    result['source'] = 'zoho_default'
    result['is_price_list_active'] = False
    
    return result


# ============================================================================
# PRICE HISTORY
# ============================================================================

async def get_price_list_history(price_list_id: int, limit: int = 100) -> Dict:
    """Get change history for price list"""
    # Verify exists
    await get_price_list_by_id(price_list_id)
    
    query = """
        SELECT 
            plh.*,
            zi.name as item_name
        FROM price_list_history plh
        LEFT JOIN zoho_items zi ON zi.id = plh.item_id
        WHERE plh.price_list_id = $1
        ORDER BY plh.changed_at DESC
        LIMIT $2
    """
    
    history = await fetch_all(query, price_list_id, limit)
    
    return {
        "history": [dict(h) for h in history],
        "total": len(history)
    }


# ============================================================================
# STATISTICS
# ============================================================================

async def get_price_list_stats() -> Dict:
    """Get price list statistics"""
    today = date.today()
    
    # Total price lists
    total = await fetch_one("SELECT COUNT(*) as count FROM customer_price_lists")
    
    # Active
    active = await fetch_one(
        """
        SELECT COUNT(*) as count 
        FROM customer_price_lists
        WHERE valid_from <= $1 
        AND (valid_to IS NULL OR valid_to >= $1)
        AND is_active = TRUE
        """,
        today
    )
    
    # Expired
    expired = await fetch_one(
        "SELECT COUNT(*) as count FROM customer_price_lists WHERE valid_to < $1",
        today
    )
    
    # Upcoming
    upcoming = await fetch_one(
        "SELECT COUNT(*) as count FROM customer_price_lists WHERE valid_from > $1",
        today
    )
    
    # Customers with price lists
    customers_with_pl = await fetch_one(
        "SELECT COUNT(DISTINCT id) as count FROM zoho_customers WHERE price_list_id IS NOT NULL"
    )
    
    # Expiring within 30 days
    from datetime import timedelta
    thirty_days = today + timedelta(days=30)
    expiring_soon = await fetch_one(
        """
        SELECT COUNT(*) as count 
        FROM customer_price_lists
        WHERE valid_to BETWEEN $1 AND $2
        AND is_active = TRUE
        """,
        today,
        thirty_days
    )
    
    return {
        "total_price_lists": total['count'] if total else 0,
        "active_price_lists": active['count'] if active else 0,
        "expired_price_lists": expired['count'] if expired else 0,
        "upcoming_price_lists": upcoming['count'] if upcoming else 0,
        "total_customers_with_price_lists": customers_with_pl['count'] if customers_with_pl else 0,
        "expiring_within_30_days": expiring_soon['count'] if expiring_soon else 0
    }
