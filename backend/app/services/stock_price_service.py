"""
================================================================================
Stock & Price Updater Service
================================================================================
Version: 1.0.0
Created: 2025-12-03

Business logic for updating WooCommerce product stock and prices with:
- List management (updatable/non-updatable/deleted)
- Bulk Excel upload/download
- Change history tracking
- Parallel WooCommerce updates using asyncio.gather()
================================================================================
"""

import asyncio
import uuid
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime, timedelta
from io import BytesIO
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import fetch_all, fetch_one, execute_query
from app.services.woocommerce_service import WooCommerceService

logger = logging.getLogger(__name__)

def safe_float(value, default=0.0):
    """Safely convert value to float, handling empty strings and None"""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


class StockPriceService:
    """Service for stock and price management"""
    
    # ========================================================================
    # Product List Management
    # ========================================================================
    
    @staticmethod
    async def get_categorized_products() -> Dict[str, List[Dict]]:
        """
        Get all products categorized into updatable/non-updatable/deleted lists
        
        Returns:
            Dict with keys: updatable, non_updatable, deleted
        """
        # Get all products from products table
        all_products = await fetch_all("""
            SELECT 
                id,
                product_id,
                variation_id,
                product_name,
                parent_product,
                sku,
                stock_quantity,
                regular_price,
                sale_price
            FROM products
            ORDER BY product_name
        """)
        
        if not all_products:
            return {"updatable": [], "non_updatable": [], "deleted": []}
        
        # Get all settings
        settings = await fetch_all("SELECT * FROM product_update_settings")
        
        # Create settings map for quick lookup
        settings_map = {}
        for setting in settings:
            key = (setting['product_id'], setting.get('variation_id'))
            settings_map[key] = setting
        
        # Categorize products
        updatable = []
        non_updatable = []
        deleted = []
        
        for product in all_products:
            key = (product['product_id'], product.get('variation_id'))
            setting = settings_map.get(key, {
                'is_updatable': True,
                'is_deleted': False,
                'notes': None
            })
            
            # Add setting info to product
            product['is_updatable'] = setting.get('is_updatable', True)
            product['is_deleted'] = setting.get('is_deleted', False)
            product['setting_notes'] = setting.get('notes')
            
            # Categorize
            if product['is_deleted']:
                deleted.append(product)
            elif product['is_updatable']:
                updatable.append(product)
            else:
                non_updatable.append(product)
        
        return {
            "updatable": updatable,
            "non_updatable": non_updatable,
            "deleted": deleted
        }
    
    # ========================================================================
    # Update Preview and Apply
    # ========================================================================
    
    @staticmethod
    async def preview_changes(changes: List[Dict]) -> Tuple[List[Dict], List[str]]:
        """
        Validate and preview changes before applying
        
        Args:
            changes: List of change dictionaries
            
        Returns:
            Tuple of (valid_changes, validation_errors)
        """
        valid_changes = []
        validation_errors = []
        
        for change in changes:
            db_id = change['db_id']
            
            # Get current product data
            product = await fetch_one("""
                SELECT id, product_id, variation_id, product_name, parent_product, sku,
                       stock_quantity, regular_price, sale_price
                FROM products
                WHERE id = $1
            """, db_id)
            
            if not product:
                validation_errors.append(f"Product ID {db_id} not found")
                continue
            
            change_details = []
            
            # Validate stock change
            if 'stock_quantity' in change and change['stock_quantity'] is not None:
                new_stock = change['stock_quantity']
                
                if new_stock < 0:
                    validation_errors.append(
                        f"{product['product_name']}: Stock cannot be negative"
                    )
                    continue
                
                if new_stock != product['stock_quantity']:
                    change_details.append({
                        'field': 'stock_quantity',
                        'old_value': product['stock_quantity'],
                        'new_value': new_stock
                    })
            
            # Validate regular price change
            if 'regular_price' in change and change['regular_price'] is not None:
                new_regular = float(change['regular_price'])
                
                if new_regular < 0:
                    validation_errors.append(
                        f"{product['product_name']}: Price cannot be negative"
                    )
                    continue
                
                if new_regular != float(product['regular_price']):
                    change_details.append({
                        'field': 'regular_price',
                        'old_value': float(product['regular_price']),
                        'new_value': new_regular
                    })
            
            # Validate sale price change
            if 'sale_price' in change and change['sale_price'] is not None:
                new_sale = float(change['sale_price'])
                
                if new_sale < 0:
                    validation_errors.append(
                        f"{product['product_name']}: Sale price cannot be negative"
                    )
                    continue
                
                # Check if sale price > regular price
                regular_price = change.get('regular_price', product['regular_price'])
                if new_sale > float(regular_price):
                    validation_errors.append(
                        f"{product['product_name']}: Sale price (Rs. {new_sale:.2f}) "
                        f"cannot be higher than regular price (Rs. {regular_price:.2f})"
                    )
                    continue
                
                if new_sale != float(product['sale_price']):
                    change_details.append({
                        'field': 'sale_price',
                        'old_value': float(product['sale_price']),
                        'new_value': new_sale
                    })
            
            # Only add if there are actual changes
            if change_details:
                valid_changes.append({
                    'db_id': db_id,
                    'product_id': product['product_id'],
                    'variation_id': product.get('variation_id'),
                    'product_name': product['product_name'],
                    'parent_product': product.get('parent_product'),
                    'sku': product.get('sku'),
                    'changes': change_details
                })
        
        return valid_changes, validation_errors
    
    @staticmethod
    async def apply_updates(changes: List[Dict], user_id: str) -> Dict:
        """
        Apply updates to database and WooCommerce
        
        Args:
            changes: List of validated changes
            user_id: User ID performing the update
            
        Returns:
            Dict with success_count, failure_count, failed_items, batch_id
        """
        batch_id = str(uuid.uuid4())
        success_count = 0
        failure_count = 0
        failed_items = []
        
        # Step 1: Update database first
        for item in changes:
            try:
                # Prepare updates
                db_updates = {}
                for change_detail in item['changes']:
                    db_updates[change_detail['field']] = change_detail['new_value']
                
                # Update product
                await execute_query("""
                    UPDATE products
                    SET stock_quantity = COALESCE($1, stock_quantity),
                        regular_price = COALESCE($2, regular_price),
                        sale_price = COALESCE($3, sale_price),
                        updated_at = NOW()
                    WHERE id = $4
                """, 
                    db_updates.get('stock_quantity'),
                    db_updates.get('regular_price'),
                    db_updates.get('sale_price'),
                    item['db_id']
                )
                
                # Log each change in history
                for change_detail in item['changes']:
                    await execute_query("""
                        INSERT INTO stock_price_history (
                            product_id, variation_id, field_changed,
                            old_value, new_value, changed_by, batch_id,
                            change_source, sync_status
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                    """,
                        item['product_id'],
                        item['variation_id'],
                        change_detail['field'],
                        str(change_detail['old_value']),
                        str(change_detail['new_value']),
                        user_id,
                        batch_id,
                        'manual',
                        'pending'
                    )
            except Exception as e:
                logger.error(f"Database update failed for {item['product_name']}: {e}")
                failed_items.append(f"{item['product_name']}: Database error - {str(e)}")
                failure_count += 1
                continue
        
        # Step 2: Update WooCommerce in parallel (using asyncio.gather)
        wc_tasks = []
        for item in changes:
            # Prepare WooCommerce updates
            wc_updates = {}
            for change_detail in item['changes']:
                field = change_detail['field']
                value = change_detail['new_value']
                
                # Format for WooCommerce API
                if field in ['regular_price', 'sale_price']:
                    wc_updates[field] = str(value)
                elif field == 'stock_quantity':
                    wc_updates[field] = int(value)
            
            # Create async task
            wc_tasks.append(
                StockPriceService._update_woocommerce_product(
                    item['product_id'],
                    item['variation_id'],
                    wc_updates,
                    batch_id,
                    item['product_name']
                )
            )
        
        # Execute WooCommerce updates concurrently (max 5 at a time)
        for i in range(0, len(wc_tasks), 5):
            batch = wc_tasks[i:i+5]
            results = await asyncio.gather(*batch, return_exceptions=True)
            
            for result in results:
                if isinstance(result, Exception):
                    failure_count += 1
                    failed_items.append(str(result))
                elif result['success']:
                    success_count += 1
                else:
                    failure_count += 1
                    failed_items.append(result['error'])
        
        return {
            'success_count': success_count,
            'failure_count': failure_count,
            'failed_items': failed_items,
            'batch_id': batch_id
        }
    
    @staticmethod
    async def _update_woocommerce_product(
        product_id: int,
        variation_id: Optional[int],
        updates: Dict,
        batch_id: str,
        product_name: str
    ) -> Dict:
        """Update a single product on WooCommerce"""
        try:
            if variation_id:
                success = await WooCommerceService.update_product_variation(
                    product_id, variation_id, updates
                )
            else:
                success = await WooCommerceService.update_product(product_id, updates)
            
            # Mark as synced
            await execute_query("""
                UPDATE stock_price_history
                SET sync_status = $1,
                    sync_attempted_at = NOW()
                WHERE batch_id = $2
                  AND product_id = $3
                  AND (variation_id = $4 OR (variation_id IS NULL AND $4 IS NULL))
            """, 
                'success' if success else 'failed',
                batch_id,
                product_id,
                variation_id
            )
            
            return {'success': success, 'error': None if success else 'WooCommerce update failed'}
        
        except Exception as e:
            error_msg = f"{product_name}: {str(e)}"
            
            # Mark as failed
            await execute_query("""
                UPDATE stock_price_history
                SET sync_status = 'failed',
                    sync_error = $1,
                    sync_attempted_at = NOW()
                WHERE batch_id = $2
                  AND product_id = $3
                  AND (variation_id = $4 OR (variation_id IS NULL AND $4 IS NULL))
            """, 
                str(e),
                batch_id,
                product_id,
                variation_id
            )
            
            return {'success': False, 'error': error_msg}
    
    # ========================================================================
    # Sync from WooCommerce
    # ========================================================================
    
    @staticmethod
    async def sync_from_woocommerce(user_id: str) -> Dict:
        """
        Sync products from WooCommerce (batch fetch and parallel updates)
        
        Returns:
            Dict with updated_count, deleted_count, unchanged_count, total_products
        """
        # Fetch all products from WooCommerce
        wc_products = await WooCommerceService.fetch_all_products()
        
        if not wc_products:
            return {
                'updated_count': 0,
                'deleted_count': 0,
                'unchanged_count': 0,
                'total_products': 0
            }
        
        # Get local products
        local_products = await fetch_all("""
            SELECT id, product_id, variation_id, stock_quantity, regular_price, sale_price
            FROM products
        """)
        
        # Create lookup for WooCommerce products
        wc_lookup = {}
        for wc_prod in wc_products:
            key = (wc_prod['id'], wc_prod.get('variation_id'))
            wc_lookup[key] = wc_prod
        
        # Identify updates and deletes
        updates_needed = []
        deletes_needed = []
        
        for local_prod in local_products:
            key = (local_prod['product_id'], local_prod.get('variation_id'))
            
            if key in wc_lookup:
                wc_prod = wc_lookup[key]
                
                # Check if values changed
                if (local_prod['stock_quantity'] != wc_prod.get('stock_quantity', 0) or
                    safe_float(local_prod['regular_price']) != safe_float(wc_prod.get('regular_price', 0)) or
                    safe_float(local_prod['sale_price']) != safe_float(wc_prod.get('sale_price', 0))):
                    
                    updates_needed.append({
                        'id': local_prod['id'],
                        'product_id': local_prod['product_id'],
                        'variation_id': local_prod.get('variation_id'),
                        'stock_quantity': wc_prod.get('stock_quantity', 0),
                        'regular_price': safe_float(wc_prod.get('regular_price', 0)),
                        'sale_price': safe_float(wc_prod.get('sale_price', 0))
                    })
            else:
                # Product not found in WooCommerce - mark as deleted
                deletes_needed.append({
                    'product_id': local_prod['product_id'],
                    'variation_id': local_prod.get('variation_id')
                })
        
        # Apply updates
        updated_count = 0
        for update in updates_needed:
            try:
                await execute_query("""
                    UPDATE products
                    SET stock_quantity = $1,
                        regular_price = $2,
                        sale_price = $3,
                        updated_at = NOW()
                    WHERE id = $4
                """, 
                    update['stock_quantity'],
                    update['regular_price'],
                    update['sale_price'],
                    update['id']
                )
                updated_count += 1
            except Exception as e:
                logger.error(f"Sync update failed: {e}")
        
        # Mark deleted products
        deleted_count = 0
        for delete in deletes_needed:
            try:
                await execute_query("""
                    INSERT INTO product_update_settings (
                        product_id, variation_id, is_deleted, updated_by
                    ) VALUES ($1, $2, true, $3)
                    ON CONFLICT (product_id, variation_id)
                    DO UPDATE SET is_deleted = true, updated_by = $3, updated_at = NOW()
                """, 
                    delete['product_id'],
                    delete['variation_id'],
                    user_id
                )
                deleted_count += 1
            except Exception as e:
                logger.error(f"Mark deleted failed: {e}")
        
        unchanged_count = len(local_products) - updated_count - deleted_count
        
        return {
            'updated_count': updated_count,
            'deleted_count': deleted_count,
            'unchanged_count': unchanged_count,
            'total_products': len(local_products)
        }
    
    # ========================================================================
    # Excel Template and Upload
    # ========================================================================
    
    @staticmethod
    async def generate_excel_template() -> bytes:
        """Generate Excel template with current updatable products"""
        products = await StockPriceService.get_categorized_products()
        updatable = products['updatable']
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Headers
        headers = [
            'product_id', 'variation_id', 'product_name', 'sku',
            'current_stock', 'current_regular_price', 'current_sale_price',
            'new_stock', 'new_regular_price', 'new_sale_price'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_num, product in enumerate(updatable, 2):
            # Combine parent and product name for variations
            display_name = product['product_name']
            if product.get('parent_product') and product.get('variation_id'):
                display_name = f"{product['parent_product']} - {product['product_name']}"
            
            ws.cell(row=row_num, column=1, value=product['product_id'])
            ws.cell(row=row_num, column=2, value=product.get('variation_id'))
            ws.cell(row=row_num, column=3, value=display_name)
            ws.cell(row=row_num, column=4, value=product.get('sku'))
            ws.cell(row=row_num, column=5, value=product['stock_quantity'])
            ws.cell(row=row_num, column=6, value=float(product['regular_price']))
            ws.cell(row=row_num, column=7, value=float(product['sale_price']))
        
        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            "1. Fill in the 'new_stock', 'new_regular_price', or 'new_sale_price' columns",
            "2. Leave blank if you do not want to update that field",
            "3. Do NOT modify the product_id, variation_id, product_name, or sku columns",
            "4. Stock must be >= 0",
            "5. Sale price cannot be higher than regular price",
            "6. Save and upload the file back to the module"
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_inst.cell(row=row_num, column=1, value=instruction)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    @staticmethod
    async def process_excel_upload(file_content: bytes, user_id: str) -> Dict:
        """Process uploaded Excel file and apply changes"""
        try:
            df = pd.read_excel(BytesIO(file_content), sheet_name='Products')
            
            # Validate columns
            required_cols = ['product_id', 'variation_id', 'new_stock', 'new_regular_price', 'new_sale_price']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                return {
                    'success': False,
                    'error': f"Missing columns: {', '.join(missing_cols)}"
                }
            
            # Parse changes
            changes = []
            for _, row in df.iterrows():
                if pd.notna(row['new_stock']) or pd.notna(row['new_regular_price']) or pd.notna(row['new_sale_price']):
                    # Get db_id
                    product = await fetch_one("""
                        SELECT id FROM products
                        WHERE product_id = $1
                          AND (variation_id = $2 OR (variation_id IS NULL AND $2 IS NULL))
                    """, 
                        int(row['product_id']),
                        int(row['variation_id']) if pd.notna(row['variation_id']) else None
                    )
                    
                    if product:
                        change = {'db_id': product['id']}
                        
                        if pd.notna(row['new_stock']):
                            change['stock_quantity'] = int(row['new_stock'])
                        if pd.notna(row['new_regular_price']):
                            change['regular_price'] = float(row['new_regular_price'])
                        if pd.notna(row['new_sale_price']):
                            change['sale_price'] = float(row['new_sale_price'])
                        
                        changes.append(change)
            
            if not changes:
                return {'success': False, 'error': 'No changes found in Excel file'}
            
            # Preview and validate
            valid_changes, errors = await StockPriceService.preview_changes(changes)
            
            if errors:
                return {'success': False, 'error': '; '.join(errors)}
            
            # Apply updates
            result = await StockPriceService.apply_updates(valid_changes, user_id)
            result['success'] = True
            return result
        
        except Exception as e:
            logger.error(f"Excel upload failed: {e}")
            return {'success': False, 'error': str(e)}
    
    # ========================================================================
    # Product Settings Management
    # ========================================================================
    
    @staticmethod
    async def update_product_setting(
        product_id: int,
        variation_id: Optional[int],
        is_updatable: bool,
        user_id: str,
        notes: Optional[str] = None
    ) -> bool:
        """Update product setting (lock/unlock)"""
        try:
            await execute_query("""
                INSERT INTO product_update_settings (
                    product_id, variation_id, is_updatable, notes, updated_by
                ) VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (product_id, variation_id)
                DO UPDATE SET
                    is_updatable = $3,
                    notes = $4,
                    updated_by = $5,
                    updated_at = NOW()
            """, 
                product_id,
                variation_id,
                is_updatable,
                notes,
                user_id
            )
            return True
        except Exception as e:
            logger.error(f"Update setting failed: {e}")
            return False
    
    @staticmethod
    async def restore_deleted_product(
        product_id: int,
        variation_id: Optional[int],
        user_id: str
    ) -> bool:
        """Restore a deleted product"""
        try:
            await execute_query("""
                INSERT INTO product_update_settings (
                    product_id, variation_id, is_deleted, updated_by
                ) VALUES ($1, $2, false, $3)
                ON CONFLICT (product_id, variation_id)
                DO UPDATE SET
                    is_deleted = false,
                    updated_by = $3,
                    updated_at = NOW()
            """, 
                product_id,
                variation_id,
                user_id
            )
            return True
        except Exception as e:
            logger.error(f"Restore failed: {e}")
            return False
    
    # ========================================================================
    # Change History and Statistics
    # ========================================================================
    
    @staticmethod
    async def get_change_history(
        limit: int = 100,
        offset: int = 0,
        product_id: Optional[int] = None
    ) -> Tuple[List[Dict], int]:
        """Get change history with pagination"""
        where_clause = "WHERE h.product_id = $3" if product_id else ""
        params = [limit, offset]
        if product_id:
            params.append(product_id)
        
        items = await fetch_all(f"""
            SELECT 
                h.*,
                u.email as changed_by_email
            FROM stock_price_history h
            LEFT JOIN auth.users u ON h.changed_by = u.id
            {where_clause}
            ORDER BY h.created_at DESC
            LIMIT $1 OFFSET $2
        """, *params)
        
        count_query = f"SELECT COUNT(*) as total FROM stock_price_history h {where_clause}"
        count_params = [product_id] if product_id else []
        total = await fetch_one(count_query, *count_params)
        
        return items, total['total'] if total else 0
    
    @staticmethod
    async def get_statistics() -> Dict:
        """Get statistics for dashboard"""
        # Get all settings
        settings = await fetch_all("SELECT * FROM product_update_settings")
        
        total = len(settings)
        updatable = len([s for s in settings if s.get('is_updatable') and not s.get('is_deleted')])
        non_updatable = len([s for s in settings if not s.get('is_updatable') and not s.get('is_deleted')])
        deleted = len([s for s in settings if s.get('is_deleted')])
        
        # Get recent changes (last 24 hours)
        recent = await fetch_one("""
            SELECT COUNT(*) as count
            FROM stock_price_history
            WHERE created_at >= NOW() - INTERVAL '24 hours'
        """)
        
        return {
            'total_products': total,
            'updatable': updatable,
            'non_updatable': non_updatable,
            'deleted': deleted,
            'recent_changes': recent['count'] if recent else 0
        }
