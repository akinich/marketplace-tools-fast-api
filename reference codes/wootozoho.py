{\rtf1\ansi\ansicpg1252\cocoartf2867
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fswiss\fcharset0 Helvetica;}
{\colortbl;\red255\green255\blue255;}
{\*\expandedcolortbl;;}
\paperw11900\paperh16840\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\pard\tx720\tx1440\tx2160\tx2880\tx3600\tx4320\tx5040\tx5760\tx6480\tx7200\tx7920\tx8640\pardirnatural\partightenfactor0

\f0\fs24 \cf0 """\
woocommerce_zoho_export.py\
\
Module Key: woocommerce_zoho_export\
\
VERSION HISTORY:\
1.0.1 - Added CSV injection protection - 11/12/25\
      SECURITY:\
      - Added CSV sanitization to prevent formula injection attacks\
1.0.0 - WooCommerce to Zoho export with product mapping - 11/11/25\
KEY FUNCTIONS:\
- Fetch completed orders from WooCommerce API with pagination\
- Map product names using database (woocommerce_products table)\
- Generate invoice numbers with sequence tracking\
- Export CSV (line items) + Excel (summary) as ZIP\
- Track export history with order totals\
- Handle rate limiting and API retries\
\
Fetch completed orders from WooCommerce between dates,\
map item names using product database (woocommerce_products table),\
then export line-item CSV + summary Excel bundled into orders_export.zip.\
\
Credentials expected in Streamlit secrets:\
[woocommerce]\
api_url = "https://your-site.com/wp-json/wc/v3"\
consumer_key = "ck_xxxxx"\
consumer_secret = "cs_xxxxx"\
"""\
\
import streamlit as st\
import pandas as pd\
import requests\
from datetime import datetime, timedelta\
from utils.csv_utils import sanitize_dataframe_for_csv\
from dateutil.parser import parse\
from typing import Dict, List, Tuple, Optional\
from io import BytesIO\
from zipfile import ZipFile\
from openpyxl.utils import get_column_letter\
from openpyxl.styles import Font, Alignment\
from auth.session import SessionManager\
from config.database import ActivityLogger, Database\
import time\
\
TOOL_NAME = "WooCommerce \uc0\u8594  Zoho Export"\
\
# Constants\
MAX_PER_PAGE = 100\
REQUEST_TIMEOUT = 30\
MAX_RETRIES = 3\
RETRY_DELAY = 2\
\
# ------------------------\
# Database Functions\
# ------------------------\
\
def get_product_mapping() -> Dict[int, Dict[str, str]]:\
    """\
    Fetch product mapping from database.\
    Returns dict: \{product_id or variation_id: \{'zoho_name', 'hsn', 'usage_units'\}\}\
    Uses variation_id as key if present, otherwise product_id.\
    """\
    try:\
        db = Database.get_client()\
        response = db.table('woocommerce_products').select(\
            'product_id, variation_id, zoho_name, hsn, usage_units'\
        ).eq('is_active', True).execute()\
        \
        mapping = \{\}\
        for row in response.data:\
            # Use variation_id if exists, otherwise product_id\
            key = row['variation_id'] if row['variation_id'] else row['product_id']\
            mapping[key] = \{\
                'zoho_name': row.get('zoho_name') or '',\
                'hsn': row.get('hsn') or '',\
                'usage_units': row.get('usage_units') or ''\
            \}\
        return mapping\
    except Exception as e:\
        st.error(f"Error fetching product mapping: \{e\}")\
        ActivityLogger.log(\
            user_id=SessionManager.get_user()['id'],\
            action_type='module_error',\
            module_key='woocommerce_zoho_export',\
            description=f"Error fetching product mapping: \{e\}",\
            success=False\
        )\
        return \{\}\
\
\
def get_last_invoice_number(prefix: str) -> Optional[int]:\
    """\
    Get the last used sequence number for given prefix from export history.\
    Returns None if no history found.\
    """\
    try:\
        db = Database.get_client()\
        response = db.table('export_history').select('sequence_number').eq(\
            'invoice_prefix', prefix\
        ).order('sequence_number', desc=True).limit(1).execute()\
        \
        if response.data and len(response.data) > 0:\
            return response.data[0]['sequence_number']\
        return None\
    except Exception as e:\
        st.warning(f"Could not fetch last invoice number: \{e\}")\
        return None\
\
\
def save_export_history(orders: List[Dict], invoice_prefix: str, \
                        start_sequence: int, start_date, end_date) -> bool:\
    """\
    Save export records to history table.\
    """\
    try:\
        db = Database.get_client()\
        user_id = SessionManager.get_user()['id']\
        \
        history_records = []\
        seq = start_sequence\
        \
        for order in orders:\
            invoice_number = f"\{invoice_prefix\}\{seq:05d\}"\
            order_total = to_float(order.get("total", 0))\
            \
            # Calculate net total (after refunds)\
            refunds = order.get("refunds") or []\
            refund_total = sum(to_float(r.get("amount") or r.get("total") or 0) for r in refunds)\
            net_total = order_total - refund_total\
            \
            history_records.append(\{\
                'invoice_number': invoice_number,\
                'invoice_prefix': invoice_prefix,\
                'sequence_number': seq,\
                'order_id': order.get('id'),\
                'order_date': order.get('date_created'),\
                'customer_name': f"\{order.get('billing',\{\}).get('first_name','')\} \{order.get('billing',\{\}).get('last_name','')\}".strip(),\
                'order_total': net_total,\
                'date_range_start': start_date.strftime('%Y-%m-%d'),\
                'date_range_end': end_date.strftime('%Y-%m-%d'),\
                'total_orders_in_export': len(orders),\
                'exported_by': user_id\
            \})\
            seq += 1\
        \
        db.table('export_history').insert(history_records).execute()\
        return True\
    except Exception as e:\
        st.error(f"Error saving export history: \{e\}")\
        ActivityLogger.log(\
            user_id=SessionManager.get_user()['id'],\
            action_type='module_error',\
            module_key='woocommerce_zoho_export',\
            description=f"Error saving export history: \{e\}",\
            success=False\
        )\
        return False\
\
\
def get_export_history(start_date: Optional[datetime] = None, \
                       end_date: Optional[datetime] = None) -> pd.DataFrame:\
    """\
    Fetch export history with optional date filtering.\
    """\
    try:\
        db = Database.get_client()\
        query = db.table('export_history').select('*')\
        \
        if start_date:\
            query = query.gte('export_date', start_date.isoformat())\
        if end_date:\
            query = query.lte('export_date', end_date.isoformat())\
        \
        response = query.order('export_date', desc=True).execute()\
        \
        if response.data:\
            return pd.DataFrame(response.data)\
        return pd.DataFrame()\
    except Exception as e:\
        st.error(f"Error fetching export history: \{e\}")\
        return pd.DataFrame()\
\
\
# ------------------------\
# Utilities\
# ------------------------\
\
def to_float(x) -> float:\
    """Convert value to float, return 0.0 if invalid."""\
    try:\
        if x is None or x == "":\
            return 0.0\
        return float(x)\
    except Exception:\
        return 0.0\
\
\
def validate_invoice_prefix(prefix: str) -> bool:\
    """Validate invoice prefix format."""\
    if not prefix or len(prefix.strip()) == 0:\
        return False\
    # Add more validation if needed (e.g., no special chars)\
    return True\
\
\
def fetch_orders(api_url: str, consumer_key: str, consumer_secret: str, \
                start_iso: str, end_iso: str) -> List[Dict]:\
    """\
    Fetch orders with pagination and retry logic.\
    Returns list of orders.\
    """\
    all_orders = []\
    page = 1\
\
    while True:\
        retries = 0\
        while retries < MAX_RETRIES:\
            try:\
                resp = requests.get(\
                    f"\{api_url.rstrip('/')\}/orders",\
                    params=\{\
                        "after": start_iso,\
                        "before": end_iso,\
                        "per_page": MAX_PER_PAGE,\
                        "page": page,\
                        "status": "any",\
                        "order": "asc",\
                        "orderby": "id"\
                    \},\
                    auth=(consumer_key, consumer_secret),\
                    timeout=REQUEST_TIMEOUT\
                )\
\
                # Rate limiting\
                if resp.status_code == 429:\
                    retry_after = int(resp.headers.get('Retry-After', 60))\
                    st.warning(f"Rate limit reached. Waiting \{retry_after\} seconds...")\
                    time.sleep(retry_after)\
                    retries += 1\
                    continue\
\
                if resp.status_code != 200:\
                    st.error(f"Error fetching orders: \{resp.status_code\} - \{resp.text\}")\
                    return []\
\
                orders = resp.json()\
                if not isinstance(orders, list):\
                    st.error("Invalid response format from WooCommerce API.")\
                    return []\
\
                if not orders:\
                    return all_orders\
\
                all_orders.extend(orders)\
                page += 1\
                break  # success -> exit retry loop\
\
            except requests.exceptions.Timeout:\
                st.error("Network timeout. Please check your connection.")\
                return []\
            except requests.exceptions.ConnectionError:\
                st.error("Unable to connect to WooCommerce. Please try again.")\
                return []\
            except requests.exceptions.RequestException as e:\
                if retries < MAX_RETRIES - 1:\
                    st.warning(f"Retrying in \{RETRY_DELAY\} seconds... (Attempt \{retries + 1\}/\{MAX_RETRIES\})")\
                    time.sleep(RETRY_DELAY)\
                    retries += 1\
                else:\
                    st.error(f"Network issue - \{str(e)\}. Please try again.")\
                    return []\
            except Exception as e:\
                st.error(f"Unexpected error: \{str(e)\}")\
                return []\
\
    return all_orders\
\
\
def transform_orders_to_rows(all_orders: List[Dict], product_mapping: Dict, \
                            invoice_prefix: str, start_sequence: int) -> Tuple[List[Dict], List[Dict], List[Dict]]:\
    """\
    Transform completed orders into CSV rows and build replacements log.\
    Returns: (csv_rows, replacements_log, completed_orders)\
    """\
    all_orders.sort(key=lambda x: x.get("id", 0))\
    completed_orders = [o for o in all_orders if o.get("status", "").lower() == "completed"]\
    \
    if not completed_orders:\
        return [], [], []\
\
    csv_rows = []\
    replacements_log = []\
    sequence_number = start_sequence\
\
    for order in completed_orders:\
        order_id = order.get("id")\
        invoice_number = f"\{invoice_prefix\}\{sequence_number:05d\}"\
        sequence_number += 1\
        \
        invoice_date = parse(order.get("date_created")).strftime("%Y-%m-%d") if order.get("date_created") else ""\
        customer_name = f"\{order.get('billing',\{\}).get('first_name','')\} \{order.get('billing',\{\}).get('last_name','')\}".strip()\
        place_of_supply = order.get('billing',\{\}).get('state','')\
        currency = order.get('currency','')\
        shipping_charge = to_float(order.get('shipping_total',0))\
        entity_discount = to_float(order.get('discount_total',0))\
\
        for item in order.get("line_items", []):\
            original_item_name = item.get("name","")\
            variation_id = item.get("variation_id", 0)\
            product_id = item.get("product_id", 0)\
            \
            # Match by variation_id first, then product_id, fallback to original name\
            lookup_id = variation_id if variation_id else product_id\
            \
            if lookup_id and lookup_id in product_mapping:\
                mapping = product_mapping[lookup_id]\
                item_name_final = mapping.get("zoho_name") or original_item_name\
                hsn = mapping.get("hsn", "")\
                usage_unit = mapping.get("usage_units", "")\
                \
                # Log replacement only if zoho_name was actually used\
                if mapping.get("zoho_name"):\
                    replacements_log.append(\{\
                        "Product ID": product_id,\
                        "Variation ID": variation_id if variation_id else "\'97",\
                        "Original WooCommerce Name": original_item_name,\
                        "Replaced Zoho Name": item_name_final,\
                        "HSN": hsn,\
                        "Usage Unit": usage_unit\
                    \})\
            else:\
                # No mapping found - use original name\
                item_name_final = original_item_name\
                hsn = ""\
                usage_unit = ""\
            \
            # Try to get HSN & usage unit from meta_data as fallback (if not from DB)\
            if not hsn or not usage_unit:\
                product_meta = item.get("meta_data", []) or []\
                for meta in product_meta:\
                    key = str(meta.get("key","")).lower()\
                    if key == "hsn" and not hsn:\
                        hsn_val = meta.get("value","")\
                        hsn = "" if hsn_val is None else str(hsn_val)\
                    if key == "usage unit" and not usage_unit:\
                        usage_val = meta.get("value","")\
                        usage_unit = "" if usage_val is None else str(usage_val)\
\
            # Tax percent\
            tax_class = item.get("tax_class") or ""\
            try:\
                item_tax_pct = float(tax_class)\
            except (TypeError, ValueError):\
                item_tax_pct = 0.0\
\
            # Format HSN with leading quote to preserve leading zeros in Excel\
            hsn_formatted = f"'\{hsn\}" if hsn else ""\
\
            # Calculate undiscounted unit price from subtotal\
            # subtotal = original price \'d7 quantity (before item discounts)\
            quantity = item.get("quantity", 0)\
            subtotal = to_float(item.get("subtotal", 0))\
            unit_price = subtotal / quantity if quantity > 0 else 0.0\
\
            row = \{\
                "Invoice Number": invoice_number,\
                "PurchaseOrder": order_id,\
                "Invoice Date": invoice_date,\
                "Invoice Status": order.get("status", "").capitalize(),\
                "Customer Name": customer_name,\
                "Place of Supply": place_of_supply,\
                "Currency Code": currency,\
                "Item Name": item_name_final,\
                "HSN/SAC": hsn_formatted,\
                "Item Type": item.get("type", "goods"),\
                "Quantity": quantity,\
                "Usage unit": usage_unit,\
                "Item Price": unit_price,\
                "Is Inclusive Tax": "FALSE",\
                "Item Tax %": item_tax_pct,\
                "Discount Type": "entity_level",\
                "Is Discount Before Tax": "TRUE",\
                "Entity Discount Amount": entity_discount,\
                "Shipping Charge": shipping_charge,\
                "Item Tax Exemption Reason": "ITEM EXEMPT FROM GST",\
                "Supply Type": "Exempted",\
                "GST Treatment": "consumer"\
            \}\
            csv_rows.append(row)\
\
    return csv_rows, replacements_log, completed_orders\
\
\
def build_summary_and_order_details(completed_orders: List[Dict], invoice_prefix: str, \
                                   start_sequence: int) -> Tuple[pd.DataFrame, pd.DataFrame]:\
    """\
    Build summary metrics DataFrame and order details DataFrame.\
    """\
    all_orders_count = len(completed_orders)\
    first_order_id = completed_orders[0].get("id") if completed_orders else None\
    last_order_id = completed_orders[-1].get("id") if completed_orders else None\
    first_invoice_number = f"\{invoice_prefix\}\{start_sequence:05d\}"\
    last_invoice_number = f"\{invoice_prefix\}\{(start_sequence + len(completed_orders) - 1):05d\}" if completed_orders else None\
\
    total_revenue_by_order_total = 0.0\
    order_details_rows = []\
    seq_temp = start_sequence\
    \
    for order in completed_orders:\
        order_total = to_float(order.get("total", 0))\
        refunds = order.get("refunds") or []\
        refund_total = sum(to_float(r.get("amount") or r.get("total") or 0) for r in refunds)\
        net_total = order_total - refund_total\
        total_revenue_by_order_total += net_total\
\
        invoice_number_temp = f"\{invoice_prefix\}\{seq_temp:05d\}"\
        seq_temp += 1\
        \
        order_details_rows.append(\{\
            "Invoice Number": invoice_number_temp,\
            "Order Number": order.get("id"),\
            "Date": parse(order.get("date_created")).strftime("%Y-%m-%d") if order.get("date_created") else "",\
            "Customer Name": f"\{order.get('billing',\{\}).get('first_name','')\} \{order.get('billing',\{\}).get('last_name','')\}".strip(),\
            "Order Total": net_total\
        \})\
\
    summary_metrics = \{\
        "Metric": [\
            "Total Orders Fetched",\
            "Completed Orders",\
            "Total Revenue (Net of Refunds)",\
            "Completed Order ID Range",\
            "Invoice Number Range"\
        ],\
        "Value": [\
            all_orders_count,\
            all_orders_count,\
            total_revenue_by_order_total,\
            f"\{first_order_id\} \uc0\u8594  \{last_order_id\}" if completed_orders else "",\
            f"\{first_invoice_number\} \uc0\u8594  \{last_invoice_number\}" if completed_orders else ""\
        ]\
    \}\
    \
    summary_df = pd.DataFrame(summary_metrics)\
    order_details_df = pd.DataFrame(order_details_rows)\
    \
    # Add grand total row\
    grand_total = order_details_df["Order Total"].sum() if not order_details_df.empty else 0.0\
    grand_total_row = \{\
        "Invoice Number": "Grand Total",\
        "Order Number": "",\
        "Date": "",\
        "Customer Name": "",\
        "Order Total": grand_total\
    \}\
    \
    if not order_details_df.empty:\
        order_details_df = pd.concat([order_details_df, pd.DataFrame([grand_total_row])], ignore_index=True)\
    else:\
        order_details_df = pd.DataFrame([grand_total_row])\
\
    return summary_df, order_details_df\
\
\
def create_excel_bytes(summary_df: pd.DataFrame, order_details_df: pd.DataFrame) -> bytes:\
    """Create Excel file bytes from dataframes."""\
    output = BytesIO()\
    with pd.ExcelWriter(output, engine='openpyxl') as writer:\
        summary_df.to_excel(writer, index=False, sheet_name="Summary Metrics")\
        order_details_df.to_excel(writer, index=False, sheet_name="Order Details")\
        \
        for sheet_name in writer.sheets:\
            ws = writer.sheets[sheet_name]\
            # Header formatting\
            for cell in ws[1]:\
                cell.font = Font(bold=True)\
                cell.alignment = Alignment(horizontal="center")\
            # Column widths\
            for col in ws.columns:\
                max_length = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2\
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length\
    \
    return output.getvalue()\
\
\
def create_zip_bytes(csv_bytes: bytes, excel_bytes: bytes, start_date, end_date) -> bytes:\
    """Create ZIP file with CSV and Excel."""\
    zip_buffer = BytesIO()\
    with ZipFile(zip_buffer, "w") as zip_file:\
        zip_file.writestr(\
            f"orders_\{start_date.strftime('%Y%m%d')\}_\{end_date.strftime('%Y%m%d')\}.csv", \
            csv_bytes\
        )\
        zip_file.writestr(\
            f"summary_report_\{start_date.strftime('%Y%m%d')\}_\{end_date.strftime('%Y%m%d')\}.xlsx", \
            excel_bytes\
        )\
    zip_buffer.seek(0)\
    return zip_buffer.getvalue()\
\
\
# ------------------------\
# UI Components\
# ------------------------\
\
def show_export_tab():\
    """Main export tab UI."""\
    st.markdown("### \uc0\u55357 \u56550  Export Orders")\
    st.markdown("Fetch completed orders from WooCommerce and export to Zoho format.")\
    st.markdown("---")\
\
    # Read credentials\
    try:\
        WC_API_URL = st.secrets["woocommerce"]["api_url"]\
        WC_CONSUMER_KEY = st.secrets["woocommerce"]["consumer_key"]\
        WC_CONSUMER_SECRET = st.secrets["woocommerce"]["consumer_secret"]\
    except Exception:\
        st.error("\uc0\u9888 \u65039  WooCommerce API credentials missing from secrets!")\
        st.info("""\
        **Required secrets:**\
        ```toml\
        [woocommerce]\
        api_url = "https://your-site.com/wp-json/wc/v3"\
        consumer_key = "ck_xxxxx"\
        consumer_secret = "cs_xxxxx"\
        ```\
        """)\
        ActivityLogger.log(\
            user_id=SessionManager.get_user()['id'],\
            action_type='module_error',\
            module_key='woocommerce_zoho_export',\
            description="Missing WooCommerce credentials in secrets",\
            success=False\
        )\
        return\
\
    # Cache product mapping\
    if 'product_mapping_cache' not in st.session_state:\
        with st.spinner("Loading product database..."):\
            st.session_state.product_mapping_cache = get_product_mapping()\
    \
    product_mapping = st.session_state.product_mapping_cache\
    st.success(f"\uc0\u9989  Product database loaded: \{len(product_mapping)\} products mapped")\
\
    # Date inputs\
    col1, col2 = st.columns(2)\
    with col1:\
        start_date = st.date_input("Start Date", value=datetime.now().date() - timedelta(days=7))\
    with col2:\
        end_date = st.date_input("End Date", value=datetime.now().date())\
\
    if start_date > end_date:\
        st.error("Start date cannot be after end date.")\
        return\
\
    # Invoice prefix and sequence\
    invoice_prefix = st.text_input(\
        "Invoice Prefix", \
        value="ECHE/2526/",\
        help="Prefix for invoice numbers (e.g., ECHE/2526/)"\
    )\
    \
    if not validate_invoice_prefix(invoice_prefix):\
        st.error("Invalid invoice prefix. Please enter a valid prefix.")\
        return\
\
    # Get last used sequence for this prefix\
    last_sequence = get_last_invoice_number(invoice_prefix)\
    suggested_sequence = (last_sequence + 1) if last_sequence is not None else 1\
    \
    if last_sequence is not None:\
        st.info(f"\uc0\u55357 \u56523  Last invoice number used: **\{invoice_prefix\}\{last_sequence:05d\}**")\
    \
    start_sequence_input = st.text_input(\
        "Starting Sequence Number",\
        value=str(suggested_sequence),\
        help="Starting sequence for invoice numbers (5 digits with leading zeros)"\
    )\
\
    # Fetch button\
    fetch_button = st.button("\uc0\u55357 \u56960  Fetch & Export Orders", type="primary")\
    \
    # Logs container\
    log_container = st.empty()\
    logs = []\
\
    def append_log(msg: str, lvl: str = "info"):\
        timestamp = datetime.now().strftime("%H:%M:%S")\
        line = f"[\{timestamp\}] \{lvl.upper()\}: \{msg\}"\
        logs.append(line)\
        log_container.text_area("Process Logs", value="\\n".join(logs), height=200)\
\
    if fetch_button:\
        # Validate sequence\
        if not start_sequence_input or start_sequence_input.strip() == "":\
            st.error("Please enter a starting sequence number.")\
            return\
        \
        try:\
            start_sequence = int(start_sequence_input)\
            if start_sequence < 1:\
                raise ValueError("Sequence must be >= 1")\
        except Exception as e:\
            st.error(f"Invalid starting sequence number: \{e\}")\
            return\
\
        append_log("Starting WooCommerce export...", "info")\
        \
        ActivityLogger.log(\
            user_id=SessionManager.get_user()['id'],\
            action_type='module_use',\
            module_key='woocommerce_zoho_export',\
            description=f"Started export for \{start_date\} to \{end_date\}"\
        )\
\
        start_iso = start_date.strftime("%Y-%m-%dT00:00:00")\
        end_iso = end_date.strftime("%Y-%m-%dT23:59:59")\
\
        # Fetch orders\
        with st.spinner("Fetching orders from WooCommerce..."):\
            all_orders = fetch_orders(WC_API_URL, WC_CONSUMER_KEY, WC_CONSUMER_SECRET, start_iso, end_iso)\
        \
        append_log(f"Fetched \{len(all_orders)\} orders from WooCommerce", "info")\
\
        if not all_orders:\
            st.warning("No orders found in this date range.")\
            append_log("No orders found", "warning")\
            return\
\
        # Transform orders\
        append_log("Processing orders and mapping products...", "info")\
        csv_rows, replacements_log, completed_orders = transform_orders_to_rows(\
            all_orders, product_mapping, invoice_prefix, start_sequence\
        )\
\
        if not completed_orders:\
            st.warning("No completed orders found in this date range.")\
            append_log("No completed orders found", "warning")\
            return\
\
        append_log(f"Processed \{len(completed_orders)\} completed orders", "info")\
\
        # Display results\
        df = pd.DataFrame(csv_rows)\
        \
        st.subheader("\uc0\u55357 \u56522  Line Items Preview (first 50 rows)")\
        st.dataframe(df.head(50), width='stretch')\
\
        if replacements_log:\
            st.subheader("\uc0\u55357 \u56580  Product Name Replacements")\
            st.dataframe(pd.DataFrame(replacements_log), width='stretch')\
            append_log(f"Applied \{len(replacements_log)\} product replacements", "info")\
\
        # Summary\
        summary_df, order_details_df = build_summary_and_order_details(\
            completed_orders, invoice_prefix, start_sequence\
        )\
        \
        st.subheader("\uc0\u55357 \u56520  Summary Metrics")\
        st.dataframe(summary_df, width='stretch')\
\
        # Prepare export files\
        try:\
            append_log("Preparing export files...", "info")\
            # SECURITY: Sanitize DataFrames to prevent CSV injection\
            safe_df = sanitize_dataframe_for_csv(df)\
            safe_summary_df = sanitize_dataframe_for_csv(summary_df)\
            safe_order_details_df = sanitize_dataframe_for_csv(order_details_df)\
\
            csv_bytes = safe_df.to_csv(index=False).encode('utf-8')\
            excel_bytes = create_excel_bytes(safe_summary_df, safe_order_details_df)\
            zip_bytes = create_zip_bytes(csv_bytes, excel_bytes, start_date, end_date)\
            append_log("Export files ready", "success")\
        except Exception as e:\
            st.error(f"Error preparing exports: \{e\}")\
            append_log(f"Error: \{e\}", "error")\
            return\
\
        # Save to history\
        append_log("Saving export history...", "info")\
        if save_export_history(completed_orders, invoice_prefix, start_sequence, start_date, end_date):\
            append_log("Export history saved successfully", "success")\
        else:\
            append_log("Warning: Could not save export history", "warning")\
\
        # Success\
        st.success(f"\uc0\u9989  Export ready \'97 \{len(completed_orders)\} completed orders")\
        \
        ActivityLogger.log(\
            user_id=SessionManager.get_user()['id'],\
            action_type='module_use',\
            module_key='woocommerce_zoho_export',\
            description=f"Exported \{len(completed_orders)\} orders (\{start_date\} to \{end_date\})",\
            metadata=\{\
                'orders_exported': len(completed_orders),\
                'date_from': str(start_date),\
                'date_to': str(end_date),\
                'invoice_prefix': invoice_prefix,\
                'start_sequence': start_sequence\
            \}\
        )\
\
        # Download button\
        st.download_button(\
            label="\uc0\u55357 \u56549  Download Export (ZIP)",\
            data=zip_bytes,\
            file_name="orders_export.zip",\
            mime="application/zip",\
            type="primary"\
        )\
\
\
def show_history_tab():\
    """Export history tab UI."""\
    st.markdown("### \uc0\u55357 \u56523  Export History")\
    st.markdown("View all previous exports with order details.")\
    st.markdown("---")\
\
    # Date filters\
    col1, col2, col3 = st.columns([2, 2, 1])\
    with col1:\
        filter_start = st.date_input(\
            "From Date",\
            value=datetime.now().date() - timedelta(days=30),\
            key="history_start"\
        )\
    with col2:\
        filter_end = st.date_input(\
            "To Date",\
            value=datetime.now().date(),\
            key="history_end"\
        )\
    with col3:\
        st.write("")  # Spacing\
        st.write("")\
        refresh_btn = st.button("\uc0\u55357 \u56580  Refresh", key="refresh_history")\
\
    # Fetch history\
    history_df = get_export_history(\
        datetime.combine(filter_start, datetime.min.time()),\
        datetime.combine(filter_end, datetime.max.time())\
    )\
\
    if history_df.empty:\
        st.info("No export history found for the selected date range.")\
        return\
\
    # Display configuration\
    display_columns = [\
        'invoice_number', 'order_id', 'order_date', \
        'customer_name', 'order_total', 'export_date'\
    ]\
    \
    # Format the dataframe\
    display_df = history_df[display_columns].copy()\
    display_df.columns = [\
        'Invoice Number', 'Order Number', 'Order Date',\
        'Customer Name', 'Order Total', 'Exported On'\
    ]\
    \
    # Format dates\
    display_df['Order Date'] = pd.to_datetime(display_df['Order Date']).dt.strftime('%Y-%m-%d %H:%M')\
    display_df['Exported On'] = pd.to_datetime(display_df['Exported On']).dt.strftime('%Y-%m-%d %H:%M')\
    \
    # Format currency\
    display_df['Order Total'] = display_df['Order Total'].apply(lambda x: f"\uc0\u8377 \{x:,.2f\}")\
\
    # Statistics\
    col1, col2, col3, col4 = st.columns(4)\
    with col1:\
        st.metric("Total Exports", len(history_df['invoice_prefix'].unique() if 'invoice_prefix' in history_df.columns else history_df))\
    with col2:\
        st.metric("Total Orders", len(history_df))\
    with col3:\
        total_revenue = history_df['order_total'].sum()\
        st.metric("Total Revenue", f"\uc0\u8377 \{total_revenue:,.2f\}")\
    with col4:\
        avg_order = history_df['order_total'].mean()\
        st.metric("Avg Order Value", f"\uc0\u8377 \{avg_order:,.2f\}")\
\
    st.markdown("---")\
    \
    # Display table\
    st.dataframe(\
        display_df,\
        width='stretch',\
        height=400\
    )\
\
    # Download history\
    if st.button("\uc0\u55357 \u56549  Download History (Excel)"):\
        output = BytesIO()\
        with pd.ExcelWriter(output, engine='openpyxl') as writer:\
            display_df.to_excel(writer, index=False, sheet_name="Export History")\
        \
        st.download_button(\
            label="Download Excel File",\
            data=output.getvalue(),\
            file_name=f"export_history_\{datetime.now().strftime('%Y%m%d_%H%M%S')\}.xlsx",\
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"\
        )\
\
\
def show_how_to_use():\
    """Display how to use guide."""\
    st.markdown("### \uc0\u55357 \u56534  How to Use This Module")\
    st.markdown("---")\
    \
    with st.expander("**\uc0\u55356 \u57263  Overview**", expanded=True):\
        st.markdown("""\
        This module exports WooCommerce orders to Zoho-compatible format with:\
        - \uc0\u9989  Automatic product name mapping from database\
        - \uc0\u9989  HSN codes and usage units from product database\
        - \uc0\u9989  Invoice number sequencing with history tracking\
        - \uc0\u9989  CSV + Excel bundled in ZIP file\
        """)\
    \
    with st.expander("**\uc0\u55357 \u56523  Prerequisites**"):\
        st.markdown("""\
        1. **WooCommerce API credentials** must be configured in secrets\
        2. **Product database** (`woocommerce_products` table) must be populated\
        3. Products should have `zoho_name`, `hsn`, and `usage_units` filled\
        """)\
    \
    with st.expander("**\uc0\u55357 \u56960  Step-by-Step Guide**"):\
        st.markdown("""\
        ### Export Tab\
        \
        1. **Select Date Range**\
           - Choose start and end dates for orders\
           - Maximum 30-31 days recommended for performance\
        \
        2. **Set Invoice Details**\
           - Enter invoice prefix (e.g., "ECHE/2526/")\
           - Starting sequence auto-populates from last export\
           - Can manually override if needed (5-digit format with leading zeros)\
        \
        3. **Fetch & Export**\
           - Click "Fetch & Export Orders" button\
           - Wait for processing (check logs)\
           - Review line items and replacements\
           - Download the ZIP file\
        \
        ### What Gets Exported?\
        \
        **orders_YYYYMMDD_YYYYMMDD.csv** - Line items with:\
        - Invoice numbers\
        - Product names (mapped to Zoho names)\
        - HSN codes and usage units\
        - Quantities, prices, taxes\
        - Customer details\
        \
        **summary_report_YYYYMMDD_YYYYMMDD.xlsx** - Contains:\
        - Summary metrics (order count, revenue, ranges)\
        - Order details with totals\
        \
        ### Product Mapping\
        \
        The module automatically matches WooCommerce products to your database:\
        1. **Variation products**: Matches by `variation_id` first\
        2. **Simple products**: Matches by `product_id`\
        3. **No match**: Uses original WooCommerce name\
        \
        Replacements are logged in the "Product Name Replacements" table.\
        """)\
    \
    with st.expander("**\uc0\u55357 \u56522  History Tab**"):\
        st.markdown("""\
        - View all previous exports with filters\
        - Track invoice numbers and sequences\
        - See order totals and export dates\
        - Download history as Excel\
        - Statistics: total exports, orders, revenue\
        """)\
    \
    with st.expander("**\uc0\u9888 \u65039  Common Issues**"):\
        st.markdown("""\
        **No orders found:**\
        - Check date range (only completed orders are exported)\
        - Verify WooCommerce has completed orders in that range\
        \
        **Product mapping issues:**\
        - Ensure products exist in `woocommerce_products` table\
        - Run Product Management module sync if needed\
        - Check that `zoho_name`, `hsn`, `usage_units` are filled\
        \
        **Invoice sequence conflicts:**\
        - History tab shows last used invoice numbers\
        - Can manually override starting sequence if needed\
        - Ensure no duplicate invoice numbers\
        \
        **API errors:**\
        - Check WooCommerce API credentials in secrets\
        - Verify API permissions in WooCommerce\
        - Check for rate limiting (handled automatically)\
        """)\
    \
    with st.expander("**\uc0\u55357 \u56481  Best Practices**"):\
        st.markdown("""\
        1. **Before First Export:**\
           - Sync all products from WooCommerce\
           - Fill in Zoho names, HSN codes, usage units\
           - Set starting sequence number carefully\
        \
        2. **Regular Exports:**\
           - Export completed orders regularly (daily/weekly)\
           - Use date filters to avoid re-exporting\
           - Check product replacements for accuracy\
        \
        3. **Data Management:**\
           - Review history tab periodically\
           - Keep product database updated\
           - Monitor invoice sequence gaps\
        \
        4. **Performance:**\
           - Limit date range to 30 days for faster exports\
           - Cache is used for product mapping (refresh if products change)\
           - Large exports may take 1-2 minutes\
        """)\
\
\
# ------------------------\
# Main Entry Point\
# ------------------------\
\
def show():\
    """Streamlit entry point for the woocommerce_zoho_export module."""\
    # Module access check\
    SessionManager.require_module_access('woocommerce_zoho_export')\
\
    user = SessionManager.get_user()\
    \
    st.markdown("### \uc0\u55357 \u56550  WooCommerce \u8594  Zoho Export")\
    st.markdown("Export completed orders from WooCommerce to Zoho-compatible format with automatic product mapping.")\
    st.markdown("---")\
\
    # Tabs\
    tab1, tab2, tab3 = st.tabs(["\uc0\u55357 \u56548  Export", "\u55357 \u56523  History", "\u55357 \u56534  How to Use"])\
    \
    with tab1:\
        show_export_tab()\
    \
    with tab2:\
        show_history_tab()\
    \
    with tab3:\
        show_how_to_use()}